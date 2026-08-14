import io
import json
from typing import List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from matching.schemas import UnifiedQueryResult

def generate_inquiry_excel(
    query_brand: str,
    query_model: str,
    quantity: int,
    results: List[UnifiedQueryResult],
    query_mode: str = "model",
    query_keyword: str = "",
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "询价与比价清单"
    ws.views.sheetView[0].showGridLines = True

    # 样式定义
    font_title = Font(name="Microsoft YaHei", size=14, bold=True, color="1F2937")
    font_header = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
    font_sub_header = Font(name="Microsoft YaHei", size=11, bold=True, color="1E3A8A")
    font_data = Font(name="Microsoft YaHei", size=9, color="374151")
    font_link = Font(name="Microsoft YaHei", size=9, color="2563EB", underline="single")
    font_badge_exact = Font(name="Microsoft YaHei", size=9, bold=True, color="166534")
    font_badge_replace = Font(name="Microsoft YaHei", size=9, bold=True, color="991B1B")

    fill_header = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    fill_exact = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fill_replace = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    fill_even = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB")
    )

    # 1. 查询条件头部信息
    ws.merge_cells("A1:N1")
    query_title = query_keyword if query_mode == "keyword" else f"{query_brand} {query_model}"
    ws["A1"] = f"商品询价比价表 - {query_title.strip()}"
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30

    ws["A2"] = f"查询模式: {'关键词' if query_mode == 'keyword' else '工业品牌+完整型号'}"
    ws["C2"] = f"查询内容: {query_title.strip()}"
    ws["E2"] = f"采购数量: {quantity}"
    for cell in ["A2", "C2", "E2"]:
        ws[cell].font = Font(name="Microsoft YaHei", size=10, bold=True)

    # 2. 表头定义
    headers = [
        "序号", "匹配等级", "置信分", "平台", "商品标题", "供应商/店铺", 
        "店铺类型", "起订量", "询价单价(元)", "小计(元)", "含税", 
        "预计交期", "原装/替代", "不匹配/风险提示", "链接"
    ]

    current_row = 4
    ws.row_dimensions[current_row].height = 24
    for col_num, h in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 冻结表头
    ws.freeze_panes = "A5"

    # 按匹配等级排序 (EXACT > HIGH > POSSIBLE > REPLACEMENT > MISMATCH)
    order_map = {"EXACT": 1, "HIGH": 2, "POSSIBLE": 3, "REPLACEMENT": 4, "MISMATCH": 5, "UNKNOWN": 6}
    sorted_results = sorted(results, key=lambda x: order_map.get(x.match_level, 99))

    # 3. 填入数据行
    for idx, item in enumerate(sorted_results, 1):
        current_row += 1
        ws.row_dimensions[current_row].height = 20
        subtotal = round(item.unit_price * quantity, 2)
        tax_str = "含税" if item.tax_included else ("未税" if item.tax_included is False else "未知")

        row_data = [
            idx,
            item.match_level,
            f"{item.match_score:.0f}",
            item.platform,
            item.title,
            item.supplier_name,
            item.supplier_type or "未知",
            item.min_order_quantity,
            item.unit_price,
            subtotal,
            tax_str,
            item.delivery_time or "未知",
            item.original_or_replacement,
            "；".join(item.mismatch_reasons) if item.mismatch_reasons else "无",
            "打开商品"
        ]

        for col_num, val in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_num, value=val)
            cell.font = font_data
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            # 居中列
            if col_num in [1, 2, 3, 4, 7, 8, 11, 13]:
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # 数字格式
            if col_num in [9, 10]:
                cell.number_format = "¥#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")

            # 匹配等级高亮样式
            if col_num == 2:
                if item.match_level in ["EXACT", "HIGH"]:
                    cell.fill = fill_exact
                    cell.font = font_badge_exact
                elif item.match_level == "REPLACEMENT":
                    cell.fill = fill_replace
                    cell.font = font_badge_replace

            # 超链接
            if col_num == 15 and item.product_url:
                cell.hyperlink = item.product_url
                cell.font = font_link

    # 4. 自动列宽调节
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            # 处理中文长度
            len_val = sum(2 if ord(c) > 127 else 1 for c in val_str)
            if len_val > max_len:
                max_len = len_val
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 45)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
